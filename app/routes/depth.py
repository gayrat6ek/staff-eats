from app.db.session import SessionLocal
from typing import Union, Any
from datetime import datetime, timedelta
import pytz
from jose import jwt
from passlib.context import CryptContext
from openpyxl import load_workbook
import bcrypt
import random
import string

from sqlalchemy.orm import Session
from typing import Union, Any
from fastapi import (
    Depends,
    HTTPException,
    status,
)
import smtplib
from pydantic import ValidationError
from fastapi.security import OAuth2PasswordBearer
import xml.etree.ElementTree as ET
import os
#from schemas import user_schema
#from queries import user_query as crud
from dotenv import load_dotenv
from fastapi import (
    Depends,
    HTTPException,
    status,
)
from jose import jwt
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.db.session import SessionLocal
from app.schemas.users import UserGet
from app.crud.users import get_user_byusername

password_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
reuseable_oauth = OAuth2PasswordBearer(tokenUrl="/api/v1/login", scheme_name="JWT")
from app.core.config import settings
#from schemas import user_schema
#from queries import user_query as crud

# Dependency to get the database session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


async def get_current_user(
    token: str = Depends(reuseable_oauth), db: Session = Depends(get_db)
) -> UserGet:
    if token == settings.backend_token:
        user: Union[dict[str, Any], None] = get_user_byusername(db, 'admin')
        return user




    payload = jwt.decode(token, settings.jwt_secret_key, algorithms=[settings.jwt_algorithm])
    expire_date = payload.get("exp")
    sub = payload.get("sub")
    password = payload.get("password")


    if datetime.fromtimestamp(expire_date) < datetime.now():
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token expired",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # except (jwt.JWTError, ValidationError):
    #     raise HTTPException(
    #         status_code=status.HTTP_403_FORBIDDEN,
    #         detail="Could not validate credentials",
    #         headers={"WWW-Authenticate": "Bearer"},
    #     )


    user: Union[dict[str, Any], None] = get_user_byusername(db, sub)

    if user is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Could not find user",
        )
    return user



def format_top5meal(statistics_top5meal):
    formatted_meals = {
        meal_name: round(float(average_rating), 1)
        for meal_name, average_rating in statistics_top5meal
    }
    return formatted_meals


def format_meals_groups(meals_groups):
    formatted_meals_groups = {}
    for group_name, month, total_orders in meals_groups:
        if month not in formatted_meals_groups:
            formatted_meals_groups[month] = {}
        formatted_meals_groups[month][group_name] = total_orders
    return formatted_meals_groups


def format_meals_company(meals_company):
    formatted_meals_company = {}
    for company_name, total_order_items_amount in meals_company:
        formatted_meals_company[company_name] = total_order_items_amount
    return formatted_meals_company


from openpyxl import Workbook


def generate_excel_from_statistics(statistics, file_path):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Meal Ratings Statistics"

    # Write the headers
    headers = ["Name", "Total Ratings", "Average Rating"]
    ws.append(headers)

    # Write the data
    for meal_name, data in statistics.items():
        row = [meal_name, data['total_ratings'], data['average_rating']]
        ws.append(row)

    # Save the workbook to the specified file path
    wb.save(file_path)
    return file_path


def generate_excell_list_of_ratings(ratings, file_path):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Ratings"

    # Write the headers
    headers = ["Branch","Company", "Comment","Meal","Rating", "Date"]
    ws.append(headers)

    # Write the data

    for rating in ratings:
        row_data = []
        if rating.client.department:
            row_data.append(rating.client.department.name)
            row_data.append(rating.client.department.company.name)
        else:
            row_data.append(' ')
            row_data.append(' ')

        row = [rating.comment,rating.meal.name,rating.rating, rating.created_at.strftime("%Y-%m-%d")]

        row_data.extend(row)
        ws.append(row)

    # Save the workbook to the specified file path
    wb.save(file_path)
    return file_path



def generate_orders_excell_list(orders, file_path):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Write the headers
    headers = ["Branch","Company", "Client","Meal","Amount", "Date"]
    ws.append(headers)

    # Write the data

    for order in orders:
        row = [order.client.department.name,order.client.department.company.name, order.client.name,order.meal.name,order.amount, order.created_at.strftime("%Y-%m-%d")]
        ws.append(row)

    # Save the workbook to the specified file path
    wb.save(file_path)
    return file_path



def generate_excell_list_of_orders(order_list,groups,file_path):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    total_group_meals = {}
    for i in groups:
        total_group_meals[i.name] = 0

    # Write the headers
    headers = ["Номер заявки", "Филиал","Организация","Время поступления","Порция"]


    ws.append(headers)


    # Write the data

    for order in order_list:
        text_toadd = ''
        row = [order.id,order.client.department.name,order.client.department.company.name,order.created_at.strftime("%Y-%m-%d")]
        for i in order.orderitem:
            text_toadd+=f"{i.amount} порция(и) {i.group.name}\n"
            total_group_meals[i.group.name]+=i.amount

        row.append(text_toadd)
        ws.append(row)
    ws.append([])

    # Write the total group meals
    ws.append(["Итого по группам:"])
    for group_name, total_amount in total_group_meals.items():
        ws.append([group_name, total_amount])

    # Save the workbook to the specified file path
    wb.save(file_path)
    return file_path









